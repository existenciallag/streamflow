#!/usr/bin/env python3
"""
StreamFlow Excel Workbook Generator
Automatically creates a pre-configured Excel workbook with all data imported
"""

import pandas as pd
import sqlite3
from pathlib import Path
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("ERROR: openpyxl not installed")
    print("Install with: pip install openpyxl")
    sys.exit(1)

def create_streamflow_workbook():
    """Create Excel workbook with all data pre-loaded"""
    print("=" * 60)
    print("StreamFlow Excel Workbook Generator")
    print("=" * 60)

    # Connect to database
    db_path = Path("../db/inventory.db")
    if not db_path.exists():
        db_path = Path("db/inventory.db")
    if not db_path.exists():
        print("ERROR: Database not found")
        sys.exit(1)

    print(f"📂 Loading database: {db_path}")
    conn = sqlite3.connect(db_path)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Tables to export
    tables = [
        ("Reagents", "reagents"),
        ("Reagent_Units", "reagent_units"),
        ("General_Reagents", "general_reagents"),
        ("General_Reagent_Units", "general_reagent_units"),
        ("Brands", "brands"),
        ("Fluorochromes", "fluorochromes"),
        ("Panels", "panels"),
        ("Panel_Reagents", "panel_reagents"),
        ("Cytometers", "cytometers"),
        ("Optical_Channels", "optical_channels"),
        ("Purchase_Orders", "purchase_orders"),
        ("Purchase_Order_Items", "purchase_order_items"),
    ]

    # Import data sheets
    for sheet_name, table_name in tables:
        print(f"  📊 Creating sheet: {sheet_name}...", end=" ")
        try:
            df = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)
            ws = wb.create_sheet(sheet_name)

            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    # Format header
                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            print(f"✅ {len(df)} rows")
        except Exception as e:
            print(f"❌ Error: {e}")

    # Create Dashboard sheet
    print(f"  📊 Creating Dashboard...")
    ws_dash = wb.create_sheet("Dashboard", 0)

    # Title
    ws_dash["A1"] = "StreamFlow Inventory Manager"
    ws_dash["A1"].font = Font(size=18, bold=True, color="1F497D")
    ws_dash.merge_cells("A1:G1")

    # Metrics labels
    metrics = [
        ("A3", "📊 INVENTORY METRICS"),
        ("A5", "Total Reagents:"),
        ("A6", "Total Units:"),
        ("A7", "Units In Use:"),
        ("A8", "Units Stored:"),
        ("A9", "Units Empty:"),
        ("A10", ""),
        ("A11", "⚠️ ALERTS"),
        ("A12", "Expiring Soon (30 days):"),
        ("A13", "Expired:"),
    ]

    for cell, value in metrics:
        ws_dash[cell] = value
        if "METRICS" in value or "ALERTS" in value:
            ws_dash[cell].font = Font(bold=True, size=12)
        else:
            ws_dash[cell].font = Font(bold=True)

    # Placeholder values (will be filled by VBA)
    ws_dash["B5"] = "=(COUNTA(Reagents!A:A)-1)"
    ws_dash["B6"] = "=(COUNTA(Reagent_Units!A:A)-1)"
    ws_dash["B7"] = "0"
    ws_dash["B8"] = "0"
    ws_dash["B9"] = "0"
    ws_dash["B12"] = "0"
    ws_dash["B13"] = "0"

    # Instructions
    ws_dash["E3"] = "📋 INSTRUCTIONS"
    ws_dash["E3"].font = Font(bold=True, size=12)
    ws_dash["E4"] = "1. Enable macros when opening this file"
    ws_dash["E5"] = "2. Import VBA modules (see SETUP_INSTRUCTIONS.md)"
    ws_dash["E6"] = "3. Add menu buttons to this dashboard"
    ws_dash["E7"] = "4. Click 'Refresh Dashboard' to update metrics"
    ws_dash["E8"] = ""
    ws_dash["E9"] = "📁 VBA Modules to import:"
    ws_dash["E10"] = "  • VBA_Module_Main.bas"
    ws_dash["E11"] = "  • VBA_Module_Inventory.bas"
    ws_dash["E12"] = "  • VBA_Module_Panels.bas"
    ws_dash["E13"] = "  • VBA_Module_Alerts.bas"

    # Column widths
    ws_dash.column_dimensions["A"].width = 25
    ws_dash.column_dimensions["B"].width = 15
    ws_dash.column_dimensions["E"].width = 40

    conn.close()

    # Save workbook
    output_path = Path("StreamFlow.xlsm")
    print(f"\n💾 Saving workbook: {output_path}")
    wb.save(output_path)

    print("\n" + "=" * 60)
    print("✅ SUCCESS!")
    print("=" * 60)
    print(f"📄 File created: {output_path.absolute()}")
    print(f"📊 Sheets created: {len(wb.sheetnames)}")
    print("\n🎯 Next Steps:")
    print("1. Open StreamFlow.xlsm in Excel")
    print("2. Import VBA modules (ALT+F11 → File → Import)")
    print("3. Add menu buttons to Dashboard")
    print("4. Save and enable macros")
    print("\n📖 See SETUP_INSTRUCTIONS.md for details")
    print("=" * 60)

if __name__ == "__main__":
    create_streamflow_workbook()
